from uuid import UUID
from typing import Optional, List
from decimal import Decimal
import hmac
import hashlib
import logging
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query, Request, Header, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.auth.models import User
from app.auth.permissions import get_current_user, require_business, require_customer
from app.orders.schemas import (
    OrderCreate,
    OrderStatusUpdate,
    OrderResponse,
    OrderListResponse,
)
from app.orders import service
from app.orders.models import Order
from app.shops.models import Shop
from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter(prefix="/api/v1/orders", tags=["orders"])


# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENT SCHEMAS
# ═══════════════════════════════════════════════════════════════════════════════

class CreatePaymentOrderRequest(BaseModel):
    order_id: str


class ConfirmPaymentRequest(BaseModel):
    order_id: str
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str


class RefundRequest(BaseModel):
    order_id: str
    amount: Optional[float] = None
    reason: str = "customer_request"


# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

import uuid
import random
import string

def generate_test_id(prefix: str = "test") -> str:
    """Generate a test ID for mock payments"""
    return f"{prefix}_{uuid.uuid4().hex[:16]}"


def get_razorpay_client():
    """Get Razorpay client instance"""
    import razorpay
    if not settings.RAZORPAY_KEY_ID or not settings.RAZORPAY_KEY_SECRET:
        raise HTTPException(status_code=500, detail="Razorpay credentials not configured")
    return razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))


def verify_payment_signature(razorpay_order_id: str, razorpay_payment_id: str, razorpay_signature: str) -> bool:
    """Verify Razorpay payment signature"""
    # In test mode, accept any signature starting with "test_"
    if settings.PAYMENT_TEST_MODE:
        return razorpay_signature.startswith("test_") or razorpay_signature == "test_signature"
    
    try:
        client = get_razorpay_client()
        client.utility.verify_payment_signature({
            "razorpay_order_id": razorpay_order_id,
            "razorpay_payment_id": razorpay_payment_id,
            "razorpay_signature": razorpay_signature,
        })
        return True
    except Exception:
        return False


@router.post("", response_model=OrderResponse)
async def create_order(
    data: OrderCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_customer),
):
    order = await service.create_order(db, current_user.id, data)
    return order


@router.get("/my", response_model=OrderListResponse)
async def get_my_orders(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_customer),
):
    orders, total = await service.get_customer_orders(
        db, current_user.id, page, per_page
    )
    return OrderListResponse(
        items=orders, total=total, page=page, per_page=per_page
    )


@router.get("/shop/{shop_id}", response_model=OrderListResponse)
async def get_shop_orders(
    shop_id: UUID,
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_business),
):
    orders, total = await service.get_shop_orders(
        db, shop_id, current_user.id, status, page, per_page
    )
    return OrderListResponse(
        items=orders, total=total, page=page, per_page=per_page
    )


@router.put("/{order_id}/status", response_model=OrderResponse)
async def update_order_status(
    order_id: UUID,
    data: OrderStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_business),
):
    order = await service.update_status(
        db, order_id, current_user.id, data.status
    )
    return order


@router.post("/{order_id}/cancel", response_model=OrderResponse)
async def cancel_order(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    order = await service.cancel_order(db, order_id, current_user.id)
    return order


# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/payments/create-order")
async def create_payment_order(
    req: CreatePaymentOrderRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Create a Razorpay payment order for checkout.
    Returns order details needed for Razorpay checkout.
    
    In TEST MODE (PAYMENT_TEST_MODE=True):
    - Returns mock Razorpay order IDs
    - No actual Razorpay API calls are made
    - Use test_signature for payment confirmation
    """
    order_id = UUID(req.order_id)
    
    # Fetch the order
    result = await db.execute(
        select(Order).where(Order.id == order_id, Order.customer_id == current_user.id)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.payment_status == "paid":
        raise HTTPException(status_code=400, detail="Order is already paid")
    
    # Convert total to paise (Razorpay uses smallest currency unit)
    amount_paise = int(Decimal(str(order.total)) * 100)
    
    # TEST MODE: Return mock payment order
    if settings.PAYMENT_TEST_MODE:
        test_order_id = f"order_test_{uuid.uuid4().hex[:16]}"
        order.payment_id = test_order_id
        await db.commit()
        
        logger.info(f"[TEST MODE] Created mock payment order {test_order_id} for order {order.order_number}")
        
        return {
            "razorpay_order_id": test_order_id,
            "razorpay_key_id": "rzp_test_mock",
            "amount": amount_paise,
            "currency": "INR",
            "order_number": order.order_number,
            "order_id": str(order.id),
            "test_mode": True,
            "test_payment_id": f"pay_test_{uuid.uuid4().hex[:16]}",
            "test_signature": "test_signature",
            "instructions": "Use test_payment_id and test_signature to confirm payment in test mode",
        }
    
    # PRODUCTION MODE: Create real Razorpay order
    client = get_razorpay_client()
    
    # Create Razorpay order
    razorpay_order = client.order.create({
        "amount": amount_paise,
        "currency": "INR",
        "receipt": order.order_number,
        "notes": {
            "order_id": str(order.id),
            "customer_id": str(current_user.id),
        }
    })
    
    # Store Razorpay order ID in our order
    order.payment_id = razorpay_order["id"]
    await db.commit()
    
    logger.info(f"Created Razorpay order {razorpay_order['id']} for order {order.order_number}")
    
    return {
        "razorpay_order_id": razorpay_order["id"],
        "razorpay_key_id": settings.RAZORPAY_KEY_ID,
        "amount": amount_paise,
        "currency": "INR",
        "order_number": order.order_number,
        "order_id": str(order.id),
        "test_mode": False,
    }


@router.post("/payments/confirm")
async def confirm_payment(
    req: ConfirmPaymentRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Confirm payment after successful Razorpay checkout.
    Verifies signature and updates order status.
    """
    order_id = UUID(req.order_id)
    
    # Verify signature
    if not verify_payment_signature(req.razorpay_order_id, req.razorpay_payment_id, req.razorpay_signature):
        raise HTTPException(status_code=400, detail="Invalid payment signature")
    
    # Fetch and update order
    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.payment_id != req.razorpay_order_id:
        raise HTTPException(status_code=400, detail="Payment order mismatch")
    
    # Update order with payment details
    order.payment_status = "paid"
    order.payment_id = req.razorpay_payment_id
    order.payment_method = "razorpay"
    order.status = "confirmed"
    
    await db.commit()
    
    logger.info(f"Payment confirmed for order {order.order_number}, payment_id: {req.razorpay_payment_id}")
    
    return {
        "status": "success",
        "message": "Payment confirmed successfully",
        "order_id": str(order.id),
        "order_number": order.order_number,
        "payment_status": order.payment_status,
        "test_mode": settings.PAYMENT_TEST_MODE,
    }


@router.post("/payments/refund")
async def process_refund(
    req: RefundRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Process refund for an order.
    In TEST MODE, refunds are simulated without actual Razorpay calls.
    """
    order_id = UUID(req.order_id)
    
    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.payment_status != "paid":
        raise HTTPException(status_code=400, detail="Order is not paid")
    
    if not order.payment_id:
        raise HTTPException(status_code=400, detail="No payment ID found for this order")
    
    # Calculate refund amount in paise
    if req.amount is None:
        refund_amount = int(Decimal(str(order.total)) * 100)
    else:
        refund_amount = int(Decimal(str(req.amount)) * 100)
    
    # TEST MODE: Simulate refund
    if settings.PAYMENT_TEST_MODE:
        refund_id = f"rfnd_test_{uuid.uuid4().hex[:16]}"
        
        # Update order status
        if req.amount is None or Decimal(str(req.amount)) >= Decimal(str(order.total)):
            order.payment_status = "refunded"
            order.status = "cancelled"
        else:
            order.payment_status = "partially_refunded"
        
        await db.commit()
        
        logger.info(f"[TEST MODE] Refund simulated for order {order.order_number}, refund_id: {refund_id}")
        
        return {
            "status": "success",
            "refund_id": refund_id,
            "amount": refund_amount / 100,
            "order_number": order.order_number,
            "test_mode": True,
        }
    
    # PRODUCTION MODE: Process actual refund
    client = get_razorpay_client()
    
    # Create refund
    refund = client.payment.refund(order.payment_id, {
        "amount": refund_amount,
        "notes": {
            "reason": req.reason,
            "order_id": str(order.id),
        }
    })
    
    # Update order status
    if req.amount is None or Decimal(str(req.amount)) >= Decimal(str(order.total)):
        order.payment_status = "refunded"
        order.status = "cancelled"
    else:
        order.payment_status = "partially_refunded"
    
    await db.commit()
    
    logger.info(f"Refund processed for order {order.order_number}, refund_id: {refund['id']}")
    
    return {
        "refund_id": refund["id"],
        "amount": refund_amount / 100,
        "status": refund["status"],
    }


@router.post("/payments/webhook")
async def razorpay_webhook(
    request: Request,
    x_razorpay_signature: str = Header(None),
    db: AsyncSession = Depends(get_db),
):
    """
    Handle Razorpay webhook events.
    """
    if not x_razorpay_signature:
        raise HTTPException(status_code=400, detail="Missing signature header")
    
    # Get raw body for signature verification
    body = await request.body()
    
    # Verify signature
    expected_signature = hmac.new(
        settings.RAZORPAY_KEY_SECRET.encode(),
        body,
        hashlib.sha256
    ).hexdigest()
    
    if not hmac.compare_digest(expected_signature, x_razorpay_signature):
        raise HTTPException(status_code=400, detail="Invalid signature")
    
    # Parse payload
    payload = await request.json()
    event_type = payload.get("event")
    
    if not event_type:
        raise HTTPException(status_code=400, detail="Missing event type")
    
    logger.info(f"Processing webhook event: {event_type}")
    
    if event_type == "payment.captured":
        payment = payload.get("payment", {}).get("entity", {})
        order_id_str = payment.get("notes", {}).get("order_id")
        
        if order_id_str:
            result = await db.execute(
                select(Order).where(Order.id == UUID(order_id_str))
            )
            order = result.scalar_one_or_none()
            
            if order and order.payment_status != "paid":
                order.payment_status = "paid"
                order.payment_id = payment.get("id")
                order.status = "confirmed"
                await db.commit()
                logger.info(f"Order {order.order_number} marked as paid via webhook")
    
    elif event_type == "payment.failed":
        payment = payload.get("payment", {}).get("entity", {})
        order_id_str = payment.get("notes", {}).get("order_id")
        
        if order_id_str:
            result = await db.execute(
                select(Order).where(Order.id == UUID(order_id_str))
            )
            order = result.scalar_one_or_none()
            
            if order:
                order.payment_status = "failed"
                await db.commit()
                logger.info(f"Order {order.order_number} payment failed via webhook")
    
    return {"status": "processed"}


@router.get("/payments/status/{order_id}")
async def get_payment_status(
    order_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get payment status for an order.
    """
    result = await db.execute(
        select(Order).where(
            Order.id == UUID(order_id),
            Order.customer_id == current_user.id
        )
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return {
        "order_id": str(order.id),
        "order_number": order.order_number,
        "payment_status": order.payment_status,
        "payment_method": order.payment_method,
        "total": float(order.total) if order.total else 0,
    }


@router.get("/{order_id}", response_model=OrderResponse)
async def get_order(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get order details by ID."""
    result = await db.execute(
        select(Order).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Verify the user is either the customer or the shop owner
    if order.customer_id != current_user.id:
        from app.shops.models import Shop
        shop_result = await db.execute(select(Shop).where(Shop.id == order.shop_id))
        shop = shop_result.scalar_one_or_none()
        if not shop or shop.owner_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to view this order")
    
    return order


# ═══════════════════════════════════════════════════════════════════════════════
# PDF INVOICE GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_invoice_pdf(order: Order, shop: Shop, customer: User) -> io.BytesIO:
    """Generate a PDF invoice for an order."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, spaceAfter=6, alignment=TA_CENTER)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
    normal_style = styles['Normal']
    bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontName='Helvetica-Bold')
    
    # Header - Shop Name & Invoice Title
    elements.append(Paragraph(f"<b>{shop.name}</b>", title_style))
    if shop.address:
        elements.append(Paragraph(shop.address, ParagraphStyle('ShopAddr', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))
    if shop.phone:
        elements.append(Paragraph(f"Phone: {shop.phone}", ParagraphStyle('ShopPhone', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))
    
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("<b>TAX INVOICE</b>", ParagraphStyle('InvoiceTitle', parent=styles['Heading2'], alignment=TA_CENTER, fontSize=14)))
    elements.append(Spacer(1, 0.5*cm))
    
    # Invoice Details Table
    order_date = order.created_at.strftime("%d %b %Y, %I:%M %p") if order.created_at else "N/A"
    invoice_data = [
        ["Invoice No:", order.order_number or f"INV-{str(order.id)[:8].upper()}"],
        ["Date:", order_date],
        ["Status:", order.status.upper() if order.status else "PENDING"],
        ["Payment:", order.payment_status.upper() if order.payment_status else "PENDING"],
    ]
    
    invoice_table = Table(invoice_data, colWidths=[2.5*cm, 5*cm])
    invoice_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    # Customer Details
    customer_data = [
        ["Bill To:"],
        [customer.name or customer.phone or "Customer"],
        [customer.phone or ""],
    ]
    
    customer_table = Table(customer_data, colWidths=[7*cm])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    # Side by side: Invoice details and Customer details
    header_table = Table([[invoice_table, customer_table]], colWidths=[8*cm, 8*cm])
    elements.append(header_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Items Table
    items_header = ["#", "Item", "Qty", "Price", "Total"]
    items_data = [items_header]
    
    items = order.items or []
    if isinstance(items, str):
        import json
        items = json.loads(items)
    
    subtotal = Decimal("0")
    for idx, item in enumerate(items, 1):
        name = item.get("name", item.get("product_name", "Item"))
        qty = item.get("quantity", 1)
        price = Decimal(str(item.get("price", item.get("unit_price", 0))))
        total = price * qty
        subtotal += total
        items_data.append([
            str(idx),
            name[:40] + "..." if len(name) > 40 else name,
            str(qty),
            f"₹{price:.2f}",
            f"₹{total:.2f}"
        ])
    
    items_table = Table(items_data, colWidths=[0.8*cm, 8*cm, 1.5*cm, 2.5*cm, 2.5*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 0.3*cm))
    
    # Totals
    discount = Decimal(str(order.discount or 0))
    delivery_fee = Decimal(str(order.delivery_fee or 0))
    total = Decimal(str(order.total or subtotal))
    
    totals_data = [
        ["Subtotal:", f"₹{subtotal:.2f}"],
    ]
    if discount > 0:
        totals_data.append(["Discount:", f"-₹{discount:.2f}"])
    if delivery_fee > 0:
        totals_data.append(["Delivery Fee:", f"₹{delivery_fee:.2f}"])
    totals_data.append(["Grand Total:", f"₹{total:.2f}"])
    
    totals_table = Table(totals_data, colWidths=[12*cm, 3.5*cm])
    totals_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 1*cm))
    
    # Footer
    elements.append(Paragraph("Thank you for shopping with us!", ParagraphStyle('Thanks', parent=normal_style, alignment=TA_CENTER, fontSize=10, textColor=colors.grey)))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph("Powered by NearShop", ParagraphStyle('Footer', parent=normal_style, alignment=TA_CENTER, fontSize=8, textColor=colors.lightgrey)))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/{order_id}/invoice")
async def get_order_invoice(
    order_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate and download PDF invoice for an order.
    Both customer and shop owner can download.
    """
    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Verify permissions
    shop_result = await db.execute(select(Shop).where(Shop.id == order.shop_id))
    shop = shop_result.scalar_one_or_none()
    
    if order.customer_id != current_user.id:
        if not shop or shop.owner_id != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to view this invoice")
    
    # Get customer info
    customer_result = await db.execute(select(User).where(User.id == order.customer_id))
    customer = customer_result.scalar_one_or_none()
    
    # Generate PDF
    pdf_buffer = generate_invoice_pdf(order, shop, customer)
    
    filename = f"invoice_{order.order_number or str(order.id)[:8]}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# CSV/EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

class ExportOrdersRequest(BaseModel):
    start_date: Optional[str] = None  # YYYY-MM-DD
    end_date: Optional[str] = None    # YYYY-MM-DD
    status: Optional[str] = None
    format: str = "csv"  # csv or xlsx


@router.post("/shop/{shop_id}/export")
async def export_shop_orders(
    shop_id: UUID,
    req: ExportOrdersRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_business),
):
    """
    Export shop orders to CSV or Excel.
    Supports date range and status filtering.
    """
    # Verify shop ownership
    shop_result = await db.execute(select(Shop).where(Shop.id == shop_id))
    shop = shop_result.scalar_one_or_none()
    
    if not shop or shop.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Build query
    query = select(Order).where(Order.shop_id == shop_id)
    
    if req.start_date:
        start = datetime.strptime(req.start_date, "%Y-%m-%d")
        query = query.where(Order.created_at >= start)
    
    if req.end_date:
        end = datetime.strptime(req.end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        query = query.where(Order.created_at <= end)
    
    if req.status:
        query = query.where(Order.status == req.status)
    
    query = query.order_by(Order.created_at.desc())
    
    result = await db.execute(query)
    orders = result.scalars().all()
    
    # Prepare data
    rows = []
    for order in orders:
        items = order.items or []
        if isinstance(items, str):
            import json
            items = json.loads(items)
        items_str = "; ".join([f"{i.get('name', 'Item')} x{i.get('quantity', 1)}" for i in items])
        
        rows.append({
            "Order ID": order.order_number or str(order.id)[:8],
            "Date": order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            "Status": order.status or "",
            "Payment Status": order.payment_status or "",
            "Payment Method": order.payment_method or "",
            "Items": items_str,
            "Subtotal": float(order.total or 0) - float(order.delivery_fee or 0) + float(order.discount or 0),
            "Discount": float(order.discount or 0),
            "Delivery Fee": float(order.delivery_fee or 0),
            "Total": float(order.total or 0),
            "Delivery Type": order.delivery_type or "",
            "Address": order.delivery_address or "",
        })
    
    if req.format == "xlsx":
        # Generate Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Header row
        headers = list(rows[0].keys()) if rows else ["No data"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"orders_{shop.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        # Generate CSV
        import csv
        
        buffer = io.StringIO()
        if rows:
            writer = csv.DictWriter(buffer, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        else:
            buffer.write("No orders found")
        
        # Convert to bytes
        csv_bytes = io.BytesIO(buffer.getvalue().encode('utf-8'))
        
        filename = f"orders_{shop.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


# ═══════════════════════════════════════════════════════════════════════════════
# REAL-TIME ORDER TRACKING (WebSocket)
# ═══════════════════════════════════════════════════════════════════════════════

from typing import Dict, Set
from fastapi import WebSocket, WebSocketDisconnect
import json
import asyncio


class OrderTrackingManager:
    """Manages WebSocket connections for order tracking."""
    
    def __init__(self):
        # order_id -> set of connected WebSocket clients
        self.order_connections: Dict[str, Set[WebSocket]] = {}
        # shop_id -> set of connected WebSocket clients (for shop dashboard)
        self.shop_connections: Dict[str, Set[WebSocket]] = {}
    
    async def connect_order(self, order_id: str, websocket: WebSocket):
        """Connect a client to track a specific order."""
        await websocket.accept()
        if order_id not in self.order_connections:
            self.order_connections[order_id] = set()
        self.order_connections[order_id].add(websocket)
    
    async def connect_shop(self, shop_id: str, websocket: WebSocket):
        """Connect a shop owner to receive real-time order updates."""
        await websocket.accept()
        if shop_id not in self.shop_connections:
            self.shop_connections[shop_id] = set()
        self.shop_connections[shop_id].add(websocket)
    
    def disconnect_order(self, order_id: str, websocket: WebSocket):
        """Disconnect a client from order tracking."""
        if order_id in self.order_connections:
            self.order_connections[order_id].discard(websocket)
            if not self.order_connections[order_id]:
                del self.order_connections[order_id]
    
    def disconnect_shop(self, shop_id: str, websocket: WebSocket):
        """Disconnect a shop owner from updates."""
        if shop_id in self.shop_connections:
            self.shop_connections[shop_id].discard(websocket)
            if not self.shop_connections[shop_id]:
                del self.shop_connections[shop_id]
    
    async def broadcast_order_update(self, order_id: str, data: dict):
        """Send order status update to all connected clients tracking this order."""
        if order_id in self.order_connections:
            dead_connections = set()
            for websocket in self.order_connections[order_id]:
                try:
                    await websocket.send_json(data)
                except Exception:
                    dead_connections.add(websocket)
            # Clean up dead connections
            for ws in dead_connections:
                self.order_connections[order_id].discard(ws)
    
    async def broadcast_shop_update(self, shop_id: str, data: dict):
        """Send update to all connected shop owners."""
        if shop_id in self.shop_connections:
            dead_connections = set()
            for websocket in self.shop_connections[shop_id]:
                try:
                    await websocket.send_json(data)
                except Exception:
                    dead_connections.add(websocket)
            # Clean up dead connections
            for ws in dead_connections:
                self.shop_connections[shop_id].discard(ws)


# Global manager instance
order_tracking_manager = OrderTrackingManager()


async def verify_ws_token(token: str, db: AsyncSession) -> Optional[User]:
    """Verify JWT token for WebSocket connection."""
    from jose import jwt, JWTError
    from app.config import get_settings
    
    settings = get_settings()
    
    try:
        payload = jwt.decode(token, settings.JWT_PUBLIC_KEY, algorithms=["RS256"])
        user_id = payload.get("sub")
        if not user_id:
            return None
        
        result = await db.execute(select(User).where(User.id == UUID(user_id)))
        return result.scalar_one_or_none()
    except (JWTError, Exception):
        return None


@router.websocket("/ws/track/{order_id}")
async def websocket_order_tracking(
    websocket: WebSocket,
    order_id: str,
    token: str = Query(...),
):
    """
    WebSocket endpoint for real-time order tracking.
    
    Connect: ws://host/api/v1/orders/ws/track/{order_id}?token=JWT
    
    Messages sent to client:
    - {"type": "connected", "order_id": "...", "status": "..."}
    - {"type": "status_update", "order_id": "...", "status": "...", "timestamp": "..."}
    - {"type": "error", "message": "..."}
    """
    from app.core.database import get_async_session
    
    async with get_async_session() as db:
        # Verify token
        user = await verify_ws_token(token, db)
        if not user:
            await websocket.close(code=4001, reason="Invalid token")
            return
        
        # Verify order exists and user has access
        result = await db.execute(select(Order).where(Order.id == UUID(order_id)))
        order = result.scalar_one_or_none()
        
        if not order:
            await websocket.close(code=4004, reason="Order not found")
            return
        
        # Check permission (customer or shop owner)
        has_access = order.customer_id == user.id
        if not has_access:
            shop_result = await db.execute(select(Shop).where(Shop.id == order.shop_id))
            shop = shop_result.scalar_one_or_none()
            has_access = shop and shop.owner_id == user.id
        
        if not has_access:
            await websocket.close(code=4003, reason="Not authorized")
            return
        
        # Connect
        await order_tracking_manager.connect_order(order_id, websocket)
        
        try:
            # Send initial status
            await websocket.send_json({
                "type": "connected",
                "order_id": order_id,
                "order_number": order.order_number,
                "status": order.status,
                "payment_status": order.payment_status,
                "created_at": order.created_at.isoformat() if order.created_at else None,
            })
            
            # Keep connection alive and listen for messages
            while True:
                try:
                    # Wait for ping or close
                    data = await asyncio.wait_for(websocket.receive_text(), timeout=30)
                    
                    # Handle ping
                    if data == "ping":
                        await websocket.send_json({"type": "pong"})
                    
                except asyncio.TimeoutError:
                    # Send heartbeat
                    await websocket.send_json({"type": "heartbeat"})
                    
        except WebSocketDisconnect:
            pass
        finally:
            order_tracking_manager.disconnect_order(order_id, websocket)


@router.websocket("/ws/shop/{shop_id}")
async def websocket_shop_orders(
    websocket: WebSocket,
    shop_id: str,
    token: str = Query(...),
):
    """
    WebSocket endpoint for shop owners to receive real-time order notifications.
    
    Connect: ws://host/api/v1/orders/ws/shop/{shop_id}?token=JWT
    
    Messages sent to client:
    - {"type": "connected", "shop_id": "..."}
    - {"type": "new_order", "order": {...}}
    - {"type": "status_update", "order_id": "...", "status": "..."}
    """
    from app.core.database import get_async_session
    
    async with get_async_session() as db:
        # Verify token
        user = await verify_ws_token(token, db)
        if not user:
            await websocket.close(code=4001, reason="Invalid token")
            return
        
        # Verify shop ownership
        shop_result = await db.execute(select(Shop).where(Shop.id == UUID(shop_id)))
        shop = shop_result.scalar_one_or_none()
        
        if not shop or shop.owner_id != user.id:
            await websocket.close(code=4003, reason="Not authorized")
            return
        
        # Connect
        await order_tracking_manager.connect_shop(shop_id, websocket)
        
        try:
            # Send connected message
            await websocket.send_json({
                "type": "connected",
                "shop_id": shop_id,
                "shop_name": shop.name,
            })
            
            # Keep connection alive
            while True:
                try:
                    data = await asyncio.wait_for(websocket.receive_text(), timeout=30)
                    if data == "ping":
                        await websocket.send_json({"type": "pong"})
                except asyncio.TimeoutError:
                    await websocket.send_json({"type": "heartbeat"})
                    
        except WebSocketDisconnect:
            pass
        finally:
            order_tracking_manager.disconnect_shop(shop_id, websocket)


async def notify_order_status_change(order_id: str, shop_id: str, status: str, order_number: str = None):
    """
    Helper function to broadcast order status changes.
    Call this when order status is updated.
    """
    timestamp = datetime.now(timezone.utc).isoformat()
    
    # Notify order trackers
    await order_tracking_manager.broadcast_order_update(order_id, {
        "type": "status_update",
        "order_id": order_id,
        "order_number": order_number,
        "status": status,
        "timestamp": timestamp,
    })
    
    # Notify shop
    await order_tracking_manager.broadcast_shop_update(shop_id, {
        "type": "status_update",
        "order_id": order_id,
        "order_number": order_number,
        "status": status,
        "timestamp": timestamp,
    })
